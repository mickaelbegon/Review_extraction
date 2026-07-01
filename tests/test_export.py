import csv
import tempfile
import unittest
from pathlib import Path

from review_extraction.export import write_csv_summary, write_xlsx_summary
from review_extraction.models import (
    ArticleResult,
    Evidence,
    FinalAnswer,
    FinalScreeningCriterion,
    FinalScreeningResult,
    TokenUsage,
)


class ExportTests(unittest.TestCase):
    def test_write_csv_summary_flattens_answers_and_evidence(self) -> None:
        result = ArticleResult(
            article_id="paper",
            source_pdf="paper.pdf",
            usage=[
                TokenUsage(
                    step="extraction",
                    model="test-model",
                    elapsed_seconds=12.345,
                    input_tokens=1000,
                    output_tokens=200,
                    total_tokens=1200,
                    input_cost_per_million=1.0,
                    output_cost_per_million=2.0,
                    estimated_cost_usd=0.0014,
                )
            ],
            processing_seconds=14.2,
            answers=[
                FinalAnswer(
                    item_id="measurement_methods",
                    final_answer=["skin_markers_3d_optical", "skin_imu"],
                    extractor_answer=["skin_markers_3d_optical"],
                    validator_status="partial",
                    validator_answer=["skin_markers_3d_optical", "skin_imu"],
                    extractor_confidence=0.7,
                    validator_confidence=0.8,
                    final_confidence=0.68,
                    evidence=[
                        Evidence(page=3, quote="Motion was captured using optical markers.", relevance="optical system"),
                        Evidence(page=4, quote="IMUs were also used.", relevance="IMU system"),
                    ],
                    rationale_short="Two methods are supported.",
                    review_required=True,
                )
            ],
        )

        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "summary.csv"
            write_csv_summary([result], out_path)

            with out_path.open(newline="", encoding="utf-8") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["article_id"], "paper")
        self.assertEqual(row["final_answer"], "skin_markers_3d_optical; skin_imu")
        self.assertEqual(row["review_required"], "True")
        self.assertEqual(row["validator_status"], "partial")
        self.assertEqual(row["evidence_pages"], "3; 4")
        self.assertIn("optical markers", row["evidence_quotes"])
        self.assertIn("IMUs", row["evidence_quotes"])
        self.assertEqual(row["usage_input_tokens"], "1000")
        self.assertEqual(row["usage_output_tokens"], "200")
        self.assertEqual(row["usage_total_tokens"], "1200")
        self.assertEqual(row["usage_estimated_cost_usd"], "0.0014")
        self.assertEqual(row["ai_elapsed_seconds"], "12.345")
        self.assertEqual(row["processing_seconds"], "14.2")

    def test_write_csv_summary_includes_screening_only_rows(self) -> None:
        result = ArticleResult(
            article_id="excluded",
            source_pdf="excluded.pdf",
            screening=FinalScreeningResult(
                article_id="excluded",
                overall_decision="exclude",
                final_confidence=0.92,
                review_required=False,
                extraction_allowed=False,
                criteria=[],
            ),
            answers=[],
        )

        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "summary.csv"
            write_csv_summary([result], out_path)

            with out_path.open(newline="", encoding="utf-8") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["article_id"], "excluded")
        self.assertEqual(row["screening_decision"], "exclude")
        self.assertEqual(row["screening_confidence"], "0.92")
        self.assertEqual(row["screening_review_required"], "False")
        self.assertEqual(row["item_id"], "")

    def test_write_xlsx_summary_creates_workbook_sheets(self) -> None:
        result = ArticleResult(
            article_id="paper",
            source_pdf="paper.pdf",
            usage=[
                TokenUsage(
                    step="screening",
                    model="screen-model",
                    elapsed_seconds=3.5,
                    input_tokens=500,
                    output_tokens=50,
                    total_tokens=550,
                    input_cost_per_million=1.0,
                    output_cost_per_million=2.0,
                    estimated_cost_usd=0.0006,
                )
            ],
            processing_seconds=4.25,
            screening=FinalScreeningResult(
                article_id="paper",
                overall_decision="include",
                final_confidence=0.91,
                review_required=True,
                extraction_allowed=False,
                criteria=[
                    FinalScreeningCriterion(
                        criterion_id="outcome",
                        final_decision="include",
                        screener_decision="include",
                        validator_status="agree",
                        validator_decision=None,
                        screener_confidence=0.9,
                        validator_confidence=0.92,
                        final_confidence=0.91,
                        evidence=[Evidence(page=2, quote="Shoulder kinematics were assessed.", relevance="outcome")],
                        rationale_short="Shoulder outcome is present.",
                        review_required=False,
                    )
                ],
            ),
            answers=[
                FinalAnswer(
                    item_id="measurement_methods",
                    final_answer=["skin_markers_3d_optical"],
                    extractor_answer=["skin_markers_3d_optical"],
                    validator_status="agree",
                    validator_answer=None,
                    extractor_confidence=0.8,
                    validator_confidence=0.9,
                    final_confidence=0.85,
                    evidence=[Evidence(page=3, quote="Optical markers were used.", relevance="measurement")],
                    rationale_short="Optical marker system.",
                    review_required=True,
                )
            ],
        )

        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "summary.xlsx"
            write_xlsx_summary([result], out_path)

            from openpyxl import load_workbook

            workbook = load_workbook(out_path)
            self.assertEqual(set(workbook.sheetnames), {"Summary", "Screening", "Extraction", "Review required", "Usage"})
            self.assertEqual(workbook["Summary"]["A2"].value, "paper")
            self.assertEqual(workbook["Screening"]["C2"].value, "outcome")
            self.assertEqual(workbook["Extraction"]["F2"].value, "measurement_methods")
            self.assertEqual(workbook["Review required"]["A2"].value, "paper")
            self.assertEqual(workbook["Usage"]["C2"].value, "screening")
            self.assertEqual(workbook["Usage"]["E2"].value, 3.5)
            self.assertEqual(workbook["Usage"]["F2"].value, 500)
            self.assertEqual(workbook["Usage"]["M2"].value, 0.0006)


if __name__ == "__main__":
    unittest.main()
