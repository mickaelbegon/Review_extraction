import tempfile
import unittest
from pathlib import Path

from review_extraction.models import ArticleResult, Evidence, FinalAnswer, FinalScreeningResult, StudyMetadataField
from review_extraction.study_pages import load_article_results, write_study_page_workbooks


class StudyPagesTests(unittest.TestCase):
    def test_write_study_page_workbook_creates_one_sheet_per_study(self) -> None:
        result = ArticleResult(
            article_id="paper",
            source_pdf="paper.pdf",
            study_metadata=[
                StudyMetadataField(
                    field_id="study_id",
                    value="Ludewig2009",
                    confidence=0.9,
                    evidence=[Evidence(page=1, quote="Ludewig 2009.", relevance="title")],
                    rationale_short="Study ID.",
                ),
                StudyMetadataField(
                    field_id="title",
                    value="Shoulder kinematics study",
                    confidence=0.8,
                    evidence=[],
                    rationale_short="Title.",
                ),
            ],
            screening=FinalScreeningResult(
                article_id="paper",
                overall_decision="include",
                final_confidence=0.9,
                review_required=False,
                extraction_allowed=True,
                criteria=[],
            ),
            answers=[
                FinalAnswer(
                    item_id="thorax_used",
                    final_answer="yes",
                    extractor_answer="yes",
                    validator_status="agree",
                    extractor_confidence=0.8,
                    validator_confidence=0.9,
                    final_confidence=0.85,
                    evidence=[],
                    rationale_short="Thorax used.",
                    review_required=False,
                )
            ],
        )

        with tempfile.TemporaryDirectory() as tmp:
            out_paths = write_study_page_workbooks([result], Path(tmp))

            from openpyxl import load_workbook

            workbook = load_workbook(out_paths[0])
            self.assertEqual(workbook.sheetnames, ["Ludewig2009"])
            sheet = workbook["Ludewig2009"]
            self.assertEqual(sheet["A1"].value, "Ludewig2009 - study review page")
            self.assertEqual(sheet["A4"].value, "Study ID")
            self.assertEqual(sheet["B4"].value, "Ludewig2009")
            self.assertEqual(sheet["E5"].value, "thorax_used")

    def test_load_article_results_skips_index_and_screening_json(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            result = ArticleResult(article_id="paper", source_pdf="paper.pdf", answers=[])
            (root / "paper.json").write_text(result.model_dump_json(indent=2), encoding="utf-8")
            (root / "paper.screening.json").write_text("{}", encoding="utf-8")
            (root / "index.json").write_text("[]", encoding="utf-8")

            loaded = load_article_results(root)

            self.assertEqual([item.article_id for item in loaded], ["paper"])


if __name__ == "__main__":
    unittest.main()
