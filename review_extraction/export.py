from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from .models import ArticleResult

SUMMARY_COLUMNS = [
    "article_id",
    "source_pdf",
    "screening_decision",
    "screening_confidence",
    "screening_review_required",
    "item_id",
    "final_answer",
    "final_confidence",
    "review_required",
    "validator_status",
    "evidence_pages",
    "evidence_quotes",
    "usage_input_tokens",
    "usage_output_tokens",
    "usage_total_tokens",
    "usage_estimated_cost_usd",
]

SCREENING_COLUMNS = [
    "article_id",
    "source_pdf",
    "criterion_id",
    "final_decision",
    "final_confidence",
    "review_required",
    "validator_status",
    "evidence_pages",
    "evidence_quotes",
]

USAGE_COLUMNS = [
    "article_id",
    "source_pdf",
    "step",
    "model",
    "input_tokens",
    "output_tokens",
    "total_tokens",
    "input_cost_per_million",
    "output_cost_per_million",
    "estimated_cost_usd",
]


def write_csv_summary(results: list[ArticleResult], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=SUMMARY_COLUMNS)
        writer.writeheader()
        writer.writerows(summary_rows(results))


def write_xlsx_summary(results: list[ArticleResult], out_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to write XLSX summaries. Recreate the conda environment.") from exc

    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    sheets = [
        ("Summary", SUMMARY_COLUMNS, summary_rows(results)),
        ("Screening", SCREENING_COLUMNS, screening_rows(results)),
        ("Extraction", SUMMARY_COLUMNS, extraction_rows(results)),
        ("Review required", SUMMARY_COLUMNS, review_required_rows(results)),
        ("Usage", USAGE_COLUMNS, usage_rows(results)),
    ]

    for sheet_name, columns, rows in sheets:
        sheet = workbook.create_sheet(sheet_name)
        _write_sheet(sheet, columns, rows)
        _style_sheet(sheet, columns, get_column_letter, Font, PatternFill, Alignment)

    workbook.save(out_path)


def summary_rows(results: list[ArticleResult]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for result in results:
        if not result.answers:
            rows.append(_summary_base_row(result))
        for answer in result.answers:
            row = _summary_base_row(result)
            row.update(
                {
                    "item_id": answer.item_id,
                    "final_answer": _stringify_answer(answer.final_answer),
                    "final_confidence": answer.final_confidence,
                    "review_required": answer.review_required,
                    "validator_status": answer.validator_status,
                    "evidence_pages": "; ".join(str(e.page) for e in answer.evidence if e.page is not None),
                    "evidence_quotes": " || ".join(e.quote for e in answer.evidence),
                }
            )
            rows.append(row)
    return rows


def screening_rows(results: list[ArticleResult]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for result in results:
        if result.screening is None:
            continue
        for criterion in result.screening.criteria:
            rows.append(
                {
                    "article_id": result.article_id,
                    "source_pdf": result.source_pdf,
                    "criterion_id": criterion.criterion_id,
                    "final_decision": criterion.final_decision,
                    "final_confidence": criterion.final_confidence,
                    "review_required": criterion.review_required,
                    "validator_status": criterion.validator_status,
                    "evidence_pages": "; ".join(str(e.page) for e in criterion.evidence if e.page is not None),
                    "evidence_quotes": " || ".join(e.quote for e in criterion.evidence),
                }
            )
    return rows


def extraction_rows(results: list[ArticleResult]) -> list[dict[str, Any]]:
    return [row for row in summary_rows(results) if row["item_id"]]


def review_required_rows(results: list[ArticleResult]) -> list[dict[str, Any]]:
    return [
        row
        for row in summary_rows(results)
        if row["review_required"] is True or row["screening_review_required"] is True
    ]


def usage_rows(results: list[ArticleResult]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for result in results:
        for usage in result.usage:
            rows.append(
                {
                    "article_id": result.article_id,
                    "source_pdf": result.source_pdf,
                    "step": usage.step,
                    "model": usage.model,
                    "input_tokens": usage.input_tokens,
                    "output_tokens": usage.output_tokens,
                    "total_tokens": usage.total_tokens,
                    "input_cost_per_million": usage.input_cost_per_million or "",
                    "output_cost_per_million": usage.output_cost_per_million or "",
                    "estimated_cost_usd": usage.estimated_cost_usd if usage.estimated_cost_usd is not None else "",
                }
            )
    return rows


def _summary_base_row(result: ArticleResult) -> dict[str, Any]:
    usage = _usage_totals(result)
    return {
        "article_id": result.article_id,
        "source_pdf": result.source_pdf,
        "screening_decision": result.screening.overall_decision if result.screening else "",
        "screening_confidence": result.screening.final_confidence if result.screening else "",
        "screening_review_required": result.screening.review_required if result.screening else "",
        "item_id": "",
        "final_answer": "",
        "final_confidence": "",
        "review_required": "",
        "validator_status": "",
        "evidence_pages": "",
        "evidence_quotes": "",
        "usage_input_tokens": usage["input_tokens"],
        "usage_output_tokens": usage["output_tokens"],
        "usage_total_tokens": usage["total_tokens"],
        "usage_estimated_cost_usd": usage["estimated_cost_usd"],
    }


def _write_sheet(sheet: Any, columns: list[str], rows: list[dict[str, Any]]) -> None:
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])


def _style_sheet(sheet: Any, columns: list[str], get_column_letter: Any, Font: Any, PatternFill: Any, Alignment: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    review_fill = PatternFill("solid", fgColor="FCE4D6")
    exclude_fill = PatternFill("solid", fgColor="F4CCCC")
    include_fill = PatternFill("solid", fgColor="D9EAD3")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    column_widths = {
        "article_id": 36,
        "source_pdf": 48,
        "screening_decision": 18,
        "screening_confidence": 18,
        "screening_review_required": 22,
        "criterion_id": 18,
        "item_id": 34,
        "final_answer": 42,
        "final_confidence": 18,
        "review_required": 18,
        "validator_status": 18,
        "evidence_pages": 16,
        "evidence_quotes": 80,
        "usage_input_tokens": 18,
        "usage_output_tokens": 18,
        "usage_total_tokens": 18,
        "usage_estimated_cost_usd": 22,
        "input_tokens": 16,
        "output_tokens": 16,
        "total_tokens": 16,
        "input_cost_per_million": 22,
        "output_cost_per_million": 22,
        "estimated_cost_usd": 18,
    }

    for index, column in enumerate(columns, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = column_widths.get(column, 20)
        for cell in sheet[letter][1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=column in {"source_pdf", "final_answer", "evidence_quotes"})
            if column.endswith("confidence") and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
            if column.endswith("cost_usd") and isinstance(cell.value, (int, float)):
                cell.number_format = "$0.000000"

    header_index = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}
    for row in sheet.iter_rows(min_row=2):
        review_value = _row_value(row, header_index, "review_required")
        screening_review_value = _row_value(row, header_index, "screening_review_required")
        screening_decision = _row_value(row, header_index, "screening_decision") or _row_value(row, header_index, "final_decision")
        fill = None
        if review_value is True or screening_review_value is True:
            fill = review_fill
        elif screening_decision == "exclude":
            fill = exclude_fill
        elif screening_decision == "include":
            fill = include_fill
        if fill is not None:
            for cell in row:
                cell.fill = fill


def _row_value(row: tuple[Any, ...], header_index: dict[str, int], column: str) -> Any:
    index = header_index.get(column)
    if index is None:
        return None
    return row[index - 1].value


def _stringify_answer(value: str | list[str]) -> str:
    if isinstance(value, list):
        return "; ".join(value)
    return value


def _usage_totals(result: ArticleResult) -> dict[str, Any]:
    input_tokens = sum(usage.input_tokens for usage in result.usage)
    output_tokens = sum(usage.output_tokens for usage in result.usage)
    total_tokens = sum(usage.total_tokens for usage in result.usage)
    costs = [usage.estimated_cost_usd for usage in result.usage if usage.estimated_cost_usd is not None]
    return {
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": total_tokens,
        "estimated_cost_usd": round(sum(costs), 6) if costs else "",
    }
