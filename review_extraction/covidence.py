from __future__ import annotations

import csv
import re
import shutil
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from .models import ArticleResult


PDF_MANIFEST_COLUMNS = ["article_id", "source_path", "target_pdf", "status"]
SCREENING_COLUMNS = [
    "article_id",
    "source_pdf",
    "covidence_decision",
    "confidence",
    "review_required",
    "extraction_allowed",
    "criteria_summary",
    "evidence_pages",
    "evidence_quotes",
]
EXTRACTION_COLUMNS = [
    "article_id",
    "source_pdf",
    "item_id",
    "answer",
    "confidence",
    "review_required",
    "validator_status",
    "evidence_pages",
    "evidence_quotes",
]


@dataclass(frozen=True)
class CovidenceImportSummary:
    copied: int
    skipped: int
    manifest_path: Path


@dataclass(frozen=True)
class CovidenceExportSummary:
    articles: int
    screening_csv: Path
    extraction_csv: Path
    workbook_path: Path


def import_pdfs(source: Path, out_dir: Path, *, manifest_path: Path | None = None) -> CovidenceImportSummary:
    out_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = manifest_path or out_dir / "covidence_pdf_manifest.csv"
    rows: list[dict[str, Any]] = []
    copied = 0
    skipped = 0

    with _source_directory(source) as source_dir:
        for pdf_path in sorted(source_dir.rglob("*.pdf")):
            article_id = _article_id_from_pdf(pdf_path)
            target_path = _unique_pdf_path(out_dir, article_id)
            status = "copied"
            if target_path.exists():
                status = "skipped_existing"
                skipped += 1
            else:
                shutil.copy2(pdf_path, target_path)
                copied += 1
            rows.append(
                {
                    "article_id": article_id,
                    "source_path": str(pdf_path),
                    "target_pdf": str(target_path),
                    "status": status,
                }
            )

    _write_csv(manifest_path, PDF_MANIFEST_COLUMNS, rows)
    return CovidenceImportSummary(copied=copied, skipped=skipped, manifest_path=manifest_path)


def export_results(results_dir: Path, out_dir: Path) -> CovidenceExportSummary:
    out_dir.mkdir(parents=True, exist_ok=True)
    results = load_results(results_dir)
    screening_rows = covidence_screening_rows(results)
    extraction_rows = covidence_extraction_rows(results)
    screening_csv = out_dir / "covidence_screening_results.csv"
    extraction_csv = out_dir / "covidence_extraction_results.csv"
    workbook_path = out_dir / "covidence_results.xlsx"

    _write_csv(screening_csv, SCREENING_COLUMNS, screening_rows)
    _write_csv(extraction_csv, EXTRACTION_COLUMNS, extraction_rows)
    _write_xlsx(workbook_path, screening_rows, extraction_rows)
    return CovidenceExportSummary(
        articles=len(results),
        screening_csv=screening_csv,
        extraction_csv=extraction_csv,
        workbook_path=workbook_path,
    )


def load_results(results_dir: Path) -> list[ArticleResult]:
    results: list[ArticleResult] = []
    for json_path in sorted(results_dir.glob("*.json")):
        if json_path.name.endswith(".screening.json") or json_path.name == "index.json":
            continue
        results.append(ArticleResult.model_validate_json(json_path.read_text(encoding="utf-8")))
    return results


def covidence_screening_rows(results: list[ArticleResult]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for result in results:
        screening = result.screening
        if screening is None:
            continue
        criteria_summary = []
        evidence_pages = []
        evidence_quotes = []
        for criterion in screening.criteria:
            criteria_summary.append(
                f"{criterion.criterion_id}: {criterion.final_decision} "
                f"(confidence={criterion.final_confidence}, review_required={criterion.review_required})"
            )
            evidence_pages.extend(str(evidence.page) for evidence in criterion.evidence if evidence.page is not None)
            evidence_quotes.extend(evidence.quote for evidence in criterion.evidence if evidence.quote)
        rows.append(
            {
                "article_id": result.article_id,
                "source_pdf": result.source_pdf,
                "covidence_decision": _covidence_decision(screening.overall_decision, screening.review_required),
                "confidence": screening.final_confidence,
                "review_required": screening.review_required,
                "extraction_allowed": screening.extraction_allowed,
                "criteria_summary": " | ".join(criteria_summary),
                "evidence_pages": "; ".join(evidence_pages),
                "evidence_quotes": " || ".join(evidence_quotes),
            }
        )
    return rows


def covidence_extraction_rows(results: list[ArticleResult]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for result in results:
        for answer in result.answers:
            rows.append(
                {
                    "article_id": result.article_id,
                    "source_pdf": result.source_pdf,
                    "item_id": answer.item_id,
                    "answer": _stringify(answer.final_answer),
                    "confidence": answer.final_confidence,
                    "review_required": answer.review_required,
                    "validator_status": answer.validator_status,
                    "evidence_pages": "; ".join(str(e.page) for e in answer.evidence if e.page is not None),
                    "evidence_quotes": " || ".join(e.quote for e in answer.evidence if e.quote),
                }
            )
    return rows


def _source_directory(source: Path):
    if source.is_dir():
        return _NullContext(source)
    if zipfile.is_zipfile(source):
        temp_dir = tempfile.TemporaryDirectory()
        with zipfile.ZipFile(source) as archive:
            archive.extractall(temp_dir.name)
        return _TempDirectoryContext(temp_dir)
    raise ValueError(f"Covidence source must be a directory or ZIP file: {source}")


class _NullContext:
    def __init__(self, path: Path) -> None:
        self.path = path

    def __enter__(self) -> Path:
        return self.path

    def __exit__(self, *args: object) -> None:
        return None


class _TempDirectoryContext:
    def __init__(self, temp_dir: tempfile.TemporaryDirectory[str]) -> None:
        self.temp_dir = temp_dir

    def __enter__(self) -> Path:
        return Path(self.temp_dir.name)

    def __exit__(self, *args: object) -> None:
        self.temp_dir.cleanup()


def _article_id_from_pdf(pdf_path: Path) -> str:
    return _safe_stem(pdf_path.stem)


def _unique_pdf_path(out_dir: Path, article_id: str) -> Path:
    candidate = out_dir / f"{article_id}.pdf"
    if not candidate.exists():
        return candidate
    counter = 2
    while True:
        candidate = out_dir / f"{article_id}_{counter}.pdf"
        if not candidate.exists():
            return candidate
        counter += 1


def _safe_stem(value: str) -> str:
    value = re.sub(r"[^\w.-]+", "_", value, flags=re.UNICODE).strip("._-")
    return value or "paper"


def _covidence_decision(decision: str, review_required: bool) -> str:
    if review_required or decision == "uncertain":
        return "maybe"
    if decision == "include":
        return "include"
    if decision == "exclude":
        return "exclude"
    return "maybe"


def _write_csv(path: Path, columns: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(path: Path, screening_rows: list[dict[str, Any]], extraction_rows: list[dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to write Covidence XLSX exports.") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, columns, rows in [
        ("Screening", SCREENING_COLUMNS, screening_rows),
        ("Extraction", EXTRACTION_COLUMNS, extraction_rows),
    ]:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(columns)
        for row in rows:
            sheet.append([row.get(column, "") for column in columns])
        _style_sheet(sheet, columns, get_column_letter, Font, PatternFill, Alignment)
    workbook.save(path)


def _style_sheet(sheet: Any, columns: list[str], get_column_letter: Any, Font: Any, PatternFill: Any, Alignment: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = _column_width(column)
        for cell in sheet[letter][1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=column.endswith("summary") or column.endswith("quotes"))


def _column_width(column: str) -> int:
    if column in {"article_id", "source_pdf"}:
        return 42
    if column.endswith("summary") or column.endswith("quotes"):
        return 80
    return min(max(len(column) + 4, 16), 28)


def _stringify(value: str | list[str]) -> str:
    if isinstance(value, list):
        return "; ".join(value)
    return value
