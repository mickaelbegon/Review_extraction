from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from .form_schema import EXTRACTION_ITEMS
from .models import ArticleResult, FinalAnswer
from .study_metadata_schema import STUDY_METADATA_ITEMS, metadata_labels
from .screening_schema import SCREENING_CRITERIA


def load_article_results(results_dir: Path) -> list[ArticleResult]:
    results: list[ArticleResult] = []
    for json_path in sorted(results_dir.glob("*.json")):
        if json_path.name == "index.json" or json_path.name.endswith(".screening.json"):
            continue
        results.append(ArticleResult.model_validate_json(json_path.read_text(encoding="utf-8")))
    return results


def write_study_page_workbooks(
    results: list[ArticleResult],
    out_dir: Path,
    *,
    split_by_letter: bool = False,
    choice_format: bool = False,
) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    if split_by_letter:
        groups: dict[str, list[ArticleResult]] = defaultdict(list)
        for result in results:
            groups[_study_letter(result)].append(result)
        prefix = "study_pages_choices" if choice_format else "study_pages"
        return [
            _write_workbook(group_results, out_dir / f"{prefix}_{letter}.xlsx", choice_format=choice_format)
            for letter, group_results in sorted(groups.items())
        ]
    filename = "study_pages_choices.xlsx" if choice_format else "study_pages.xlsx"
    return [_write_workbook(results, out_dir / filename, choice_format=choice_format)]


def _write_workbook(results: list[ArticleResult], out_path: Path, *, choice_format: bool = False) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to write study page workbooks. Recreate the conda environment.") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    if not results:
        sheet = workbook.create_sheet("No studies")
        sheet["A1"] = "No article JSON files found."
    for result in results:
        sheet = workbook.create_sheet(_sheet_title(result, workbook.sheetnames))
        if choice_format:
            _write_choice_study_sheet(sheet, result, get_column_letter, Font, PatternFill, Alignment, Border, Side)
        else:
            _write_study_sheet(sheet, result, get_column_letter, Font, PatternFill, Alignment, Border, Side)
    workbook.save(out_path)
    return out_path


def _write_choice_study_sheet(
    sheet: Any,
    result: ArticleResult,
    get_column_letter: Any,
    Font: Any,
    PatternFill: Any,
    Alignment: Any,
    Border: Any,
    Side: Any,
) -> None:
    metadata = {field.field_id: field for field in result.study_metadata}
    answers = {answer.item_id: answer for answer in result.answers}
    criteria_by_id = {criterion.id: criterion for criterion in SCREENING_CRITERIA}
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    value_fill = PatternFill("solid", fgColor="F7FBFF")
    review_fill = PatternFill("solid", fgColor="FCE4D6")
    skipped_fill = PatternFill("solid", fgColor="E7E6E6")

    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35

    widths = [20, 38, 34, 30, 13, 14, 16, 56, 40]
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width

    sheet.merge_cells("A1:I1")
    title = _metadata_value(metadata, "study_id") or result.article_id
    sheet["A1"] = f"{title} - choice-based review sheet"
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "Section",
        "Question / field",
        "Allowed choices / coding",
        "Selected answer",
        "Confidence",
        "Review required",
        "Status",
        "Evidence used",
        "Rationale / notes",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=column)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border

    row = 4
    row = _write_choice_section(sheet, row, "1. Screening", section_fill, Font, Alignment)
    screening = result.screening
    if screening is None:
        row = _write_choice_row(
            sheet,
            row,
            "Screening",
            "Overall screening decision",
            "include / exclude / uncertain",
            "NR",
            "",
            "",
            "",
            "",
            "No screening result available.",
            skipped_fill,
            border,
        )
    else:
        row = _write_choice_row(
            sheet,
            row,
            "Screening",
            "Overall screening decision",
            "include / exclude / uncertain",
            screening.overall_decision,
            screening.final_confidence,
            screening.review_required,
            "extraction allowed" if screening.extraction_allowed else "extraction skipped",
            "",
            "",
            review_fill if screening.review_required or not screening.extraction_allowed else value_fill,
            border,
        )
        for criterion in screening.criteria:
            schema = criteria_by_id.get(criterion.criterion_id)
            allowed = "include / exclude / unclear"
            if schema is not None:
                allowed = f"include: {schema.include} | exclude: {schema.exclude}"
            row = _write_choice_row(
                sheet,
                row,
                f"Screening: {criterion.criterion_id}",
                schema.theme if schema is not None else criterion.criterion_id,
                allowed,
                criterion.final_decision,
                criterion.final_confidence,
                criterion.review_required,
                criterion.validator_status,
                _evidence_text(criterion.evidence),
                criterion.rationale_short,
                review_fill if criterion.review_required or criterion.final_decision != "include" else value_fill,
                border,
            )

    row += 1
    row = _write_choice_section(sheet, row, "2. Study-level data", section_fill, Font, Alignment)
    if not result.study_metadata:
        row = _write_choice_row(
            sheet,
            row,
            "Study-level data",
            "Study metadata extraction",
            "Only extracted when screening allows extraction",
            "Not extracted",
            "",
            "",
            "skipped",
            "",
            "Screening did not allow detailed extraction.",
            skipped_fill,
            border,
        )
    for item in STUDY_METADATA_ITEMS:
        field = metadata.get(item.id)
        row = _write_choice_row(
            sheet,
            row,
            "Study-level data",
            item.label,
            item.coding,
            field.value if field else "NR",
            field.confidence if field else "",
            "",
            "",
            _evidence_text(field.evidence) if field else "",
            field.rationale_short if field else item.guidance,
            value_fill if field else skipped_fill,
            border,
        )

    row += 1
    row = _write_choice_section(sheet, row, "3. Word extraction form", section_fill, Font, Alignment)
    if not result.answers:
        row = _write_choice_row(
            sheet,
            row,
            "Word extraction form",
            "Detailed methodological extraction",
            "Only extracted when screening allows extraction",
            "Not extracted",
            "",
            "",
            "skipped",
            "",
            "Screening did not allow detailed extraction.",
            skipped_fill,
            border,
        )
    for item in EXTRACTION_ITEMS:
        answer = answers.get(item.id)
        row = _write_choice_row(
            sheet,
            row,
            item.theme,
            item.question,
            " / ".join(item.allowed_answers),
            _stringify(answer.final_answer) if answer else "NR",
            answer.final_confidence if answer else "",
            answer.review_required if answer else "",
            answer.validator_status if answer else "",
            _evidence_text(answer.evidence) if answer else "",
            answer.rationale_short if answer else item.guidance,
            review_fill if answer and answer.review_required else value_fill if answer else skipped_fill,
            border,
        )

    for row_cells in sheet.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:I{sheet.max_row}"


def _write_choice_section(sheet: Any, row: int, title: str, fill: Any, Font: Any, Alignment: Any) -> int:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = sheet.cell(row=row, column=1)
    cell.value = title
    cell.font = Font(bold=True)
    cell.fill = fill
    cell.alignment = Alignment(horizontal="left")
    return row + 1


def _write_choice_row(
    sheet: Any,
    row: int,
    section: str,
    question: str,
    allowed_choices: str,
    selected_answer: Any,
    confidence: Any,
    review_required: Any,
    status: Any,
    evidence: str,
    rationale: str,
    fill: Any,
    border: Any,
) -> int:
    values = [
        section,
        question,
        allowed_choices,
        selected_answer if selected_answer not in {None, ""} else "NR",
        confidence,
        review_required,
        status,
        evidence,
        rationale,
    ]
    for column, value in enumerate(values, start=1):
        cell = sheet.cell(row=row, column=column)
        cell.value = value
        cell.fill = fill
        cell.border = border
    return row + 1


def _write_study_sheet(sheet: Any, result: ArticleResult, get_column_letter: Any, Font: Any, PatternFill: Any, Alignment: Any, Border: Any, Side: Any) -> None:
    labels = metadata_labels()
    metadata = {field.field_id: field for field in result.study_metadata}
    answers = {answer.item_id: answer for answer in result.answers}
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    value_fill = PatternFill("solid", fgColor="F7FBFF")
    review_fill = PatternFill("solid", fgColor="FCE4D6")
    auto_fill = PatternFill("solid", fgColor="E2F0D9")

    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35

    for col, width in enumerate([24, 30, 16, 34, 18, 18, 30, 18], start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width

    sheet.merge_cells("A1:H1")
    title = _metadata_value(metadata, "study_id") or result.article_id
    sheet["A1"] = f"{title} - study review page"
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(horizontal="center")

    row = 3
    row = _write_section_title(sheet, row, "Study identification and population", section_fill, Font, Alignment)
    for item in STUDY_METADATA_ITEMS:
        row = _write_key_value(
            sheet,
            row,
            item.label,
            _metadata_value(metadata, item.id),
            _metadata_confidence(metadata, item.id),
            _metadata_evidence(metadata, item.id),
            value_fill,
            border,
            Alignment,
        )

    row += 1
    row = _write_section_title(sheet, row, "Screening decision", section_fill, Font, Alignment)
    screening = result.screening
    screening_values = [
        ("Overall decision", screening.overall_decision if screening else ""),
        ("Confidence", screening.final_confidence if screening else ""),
        ("Review required", screening.review_required if screening else ""),
    ]
    for label, value in screening_values:
        row = _write_key_value(sheet, row, label, value, "", "", value_fill, border, Alignment)
    if screening:
        for criterion in screening.criteria:
            row = _write_key_value(
                sheet,
                row,
                f"Criterion: {criterion.criterion_id}",
                criterion.final_decision,
                criterion.final_confidence,
                _evidence_text(criterion.evidence),
                review_fill if criterion.review_required else value_fill,
                border,
                Alignment,
            )

    extraction_start = 3
    _write_extraction_blocks(
        sheet,
        extraction_start,
        answers,
        section_fill,
        value_fill,
        review_fill,
        auto_fill,
        border,
        Font,
        Alignment,
    )

    for row_cells in sheet.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A3"


def _write_extraction_blocks(sheet: Any, start_row: int, answers: dict[str, FinalAnswer], section_fill: Any, value_fill: Any, review_fill: Any, auto_fill: Any, border: Any, Font: Any, Alignment: Any) -> None:
    row = start_row
    sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
    cell = sheet.cell(row=row, column=5)
    cell.value = "Methodological extraction"
    cell.font = Font(bold=True)
    cell.fill = section_fill
    cell.alignment = Alignment(horizontal="center")
    row += 1
    headers = ["Item", "Answer", "Confidence", "Status"]
    for offset, header in enumerate(headers, start=5):
        cell = sheet.cell(row=row, column=offset)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = section_fill
        cell.border = border
    row += 1
    for item in EXTRACTION_ITEMS:
        answer = answers.get(item.id)
        if answer is None:
            continue
        fill = review_fill if answer.review_required else auto_fill if answer.validator_status == "auto_absent" else value_fill
        values = [
            item.id,
            _stringify(answer.final_answer),
            answer.final_confidence,
            answer.validator_status,
        ]
        for offset, value in enumerate(values, start=5):
            cell = sheet.cell(row=row, column=offset)
            cell.value = value
            cell.fill = fill
            cell.border = border
        row += 1


def _write_section_title(sheet: Any, row: int, title: str, fill: Any, Font: Any, Alignment: Any) -> int:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = sheet.cell(row=row, column=1)
    cell.value = title
    cell.font = Font(bold=True)
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center")
    return row + 1


def _write_key_value(sheet: Any, row: int, label: str, value: Any, confidence: Any, evidence: Any, fill: Any, border: Any, Alignment: Any) -> int:
    values = [label, value if value not in {None, ""} else "NR", confidence, evidence]
    for column, cell_value in enumerate(values, start=1):
        cell = sheet.cell(row=row, column=column)
        cell.value = cell_value
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    return row + 1


def _metadata_value(metadata: dict[str, Any], field_id: str) -> str:
    field = metadata.get(field_id)
    return field.value if field else ""


def _metadata_confidence(metadata: dict[str, Any], field_id: str) -> float | str:
    field = metadata.get(field_id)
    return field.confidence if field else ""


def _metadata_evidence(metadata: dict[str, Any], field_id: str) -> str:
    field = metadata.get(field_id)
    return _evidence_text(field.evidence) if field else ""


def _evidence_text(evidence: list[Any]) -> str:
    snippets = []
    for item in evidence[:2]:
        page = f"p.{item.page}: " if item.page is not None else ""
        snippets.append(f"{page}{item.quote}")
    return " | ".join(snippets)


def _stringify(value: str | list[str]) -> str:
    return "; ".join(value) if isinstance(value, list) else str(value)


def _study_letter(result: ArticleResult) -> str:
    study_id = next((field.value for field in result.study_metadata if field.field_id == "study_id" and field.value), result.article_id)
    match = re.search(r"[A-Za-z]", study_id)
    return match.group(0).upper() if match else "Other"


def _sheet_title(result: ArticleResult, existing: list[str]) -> str:
    base = next((field.value for field in result.study_metadata if field.field_id == "study_id" and field.value), result.article_id)
    base = re.sub(r"[\[\]\:\*\?\/\\]", "_", base)[:31] or "Study"
    title = base
    counter = 2
    while title in existing:
        suffix = f"_{counter}"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    return title
