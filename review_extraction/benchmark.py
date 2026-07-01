from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from .models import ArticleResult, FinalAnswer


@dataclass(frozen=True)
class ModelBenchmark:
    model: str
    article_rows: list[dict[str, Any]]
    disagreement_rows: list[dict[str, Any]]


def load_reference_results(reference_dir: Path) -> dict[str, ArticleResult]:
    results: dict[str, ArticleResult] = {}
    for json_path in sorted(reference_dir.glob("*.json")):
        if json_path.name.endswith(".screening.json") or json_path.name in {"index.json"}:
            continue
        result = ArticleResult.model_validate_json(json_path.read_text(encoding="utf-8"))
        result.article_id = json_path.stem
        results[result.article_id] = result
    return results


def compare_model_results(
    *,
    model: str,
    reference_results: dict[str, ArticleResult],
    candidate_results: Iterable[ArticleResult],
) -> ModelBenchmark:
    article_rows: list[dict[str, Any]] = []
    disagreement_rows: list[dict[str, Any]] = []
    for candidate in candidate_results:
        reference = reference_results.get(candidate.article_id)
        if reference is None:
            continue
        row, disagreements = compare_article(model=model, reference=reference, candidate=candidate)
        article_rows.append(row)
        disagreement_rows.extend(disagreements)
    return ModelBenchmark(model=model, article_rows=article_rows, disagreement_rows=disagreement_rows)


def compare_article(
    *,
    model: str,
    reference: ArticleResult,
    candidate: ArticleResult,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    ref_screening = reference.screening
    cand_screening = candidate.screening
    ref_decision = ref_screening.overall_decision if ref_screening else ""
    cand_decision = cand_screening.overall_decision if cand_screening else ""
    screening_match = ref_decision == cand_decision

    ref_answers = {answer.item_id: answer for answer in reference.answers}
    cand_answers = {answer.item_id: answer for answer in candidate.answers}
    item_ids = sorted(set(ref_answers) | set(cand_answers))
    matched_items = 0
    confidence_diffs: list[float] = []
    disagreements: list[dict[str, Any]] = []

    for item_id in item_ids:
        ref_answer = ref_answers.get(item_id)
        cand_answer = cand_answers.get(item_id)
        answers_match = ref_answer is not None and cand_answer is not None and _answer_key(ref_answer) == _answer_key(cand_answer)
        if answers_match:
            matched_items += 1
        if ref_answer is not None and cand_answer is not None:
            confidence_diffs.append(abs(float(ref_answer.final_confidence) - float(cand_answer.final_confidence)))
        if not answers_match:
            disagreements.append(
                {
                    "model": model,
                    "article_id": reference.article_id,
                    "kind": "answer",
                    "item_id": item_id,
                    "reference_value": _answer_text(ref_answer),
                    "candidate_value": _answer_text(cand_answer),
                    "reference_confidence": ref_answer.final_confidence if ref_answer else "",
                    "candidate_confidence": cand_answer.final_confidence if cand_answer else "",
                }
            )

    if not screening_match:
        disagreements.insert(
            0,
            {
                "model": model,
                "article_id": reference.article_id,
                "kind": "screening",
                "item_id": "overall_decision",
                "reference_value": ref_decision,
                "candidate_value": cand_decision,
                "reference_confidence": ref_screening.final_confidence if ref_screening else "",
                "candidate_confidence": cand_screening.final_confidence if cand_screening else "",
            },
        )

    compared_items = len(item_ids)
    input_tokens = sum(usage.input_tokens for usage in candidate.usage)
    cached_input_tokens = sum(usage.cached_input_tokens for usage in candidate.usage)
    output_tokens = sum(usage.output_tokens for usage in candidate.usage)
    total_tokens = sum(usage.total_tokens for usage in candidate.usage)
    costs = [usage.estimated_cost_usd for usage in candidate.usage if usage.estimated_cost_usd is not None]

    row = {
        "model": model,
        "article_id": reference.article_id,
        "screening_reference": ref_decision,
        "screening_candidate": cand_decision,
        "screening_match": screening_match,
        "reference_review_required": ref_screening.review_required if ref_screening else "",
        "candidate_review_required": cand_screening.review_required if cand_screening else "",
        "reference_items": len(ref_answers),
        "candidate_items": len(cand_answers),
        "compared_items": compared_items,
        "matching_items": matched_items,
        "answer_agreement_rate": round(matched_items / compared_items, 4) if compared_items else 1.0,
        "mean_confidence_abs_diff": round(sum(confidence_diffs) / len(confidence_diffs), 4) if confidence_diffs else "",
        "candidate_input_tokens": input_tokens,
        "candidate_cached_input_tokens": cached_input_tokens,
        "candidate_output_tokens": output_tokens,
        "candidate_total_tokens": total_tokens,
        "candidate_estimated_cost_usd": round(sum(costs), 6) if costs else "",
    }
    return row, disagreements


def summarize_benchmarks(benchmarks: list[ModelBenchmark]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for benchmark in benchmarks:
        article_rows = benchmark.article_rows
        if not article_rows:
            rows.append(_empty_model_summary(benchmark.model))
            continue
        screening_matches = sum(1 for row in article_rows if row["screening_match"] is True)
        compared_items = sum(int(row["compared_items"]) for row in article_rows)
        matching_items = sum(int(row["matching_items"]) for row in article_rows)
        confidence_diffs = [
            float(row["mean_confidence_abs_diff"])
            for row in article_rows
            if row["mean_confidence_abs_diff"] != ""
        ]
        total_costs = [
            float(row["candidate_estimated_cost_usd"])
            for row in article_rows
            if row["candidate_estimated_cost_usd"] != ""
        ]
        rows.append(
            {
                "model": benchmark.model,
                "articles": len(article_rows),
                "screening_matches": screening_matches,
                "screening_agreement_rate": round(screening_matches / len(article_rows), 4),
                "compared_items": compared_items,
                "matching_items": matching_items,
                "answer_agreement_rate": round(matching_items / compared_items, 4) if compared_items else 1.0,
                "mean_confidence_abs_diff": round(sum(confidence_diffs) / len(confidence_diffs), 4) if confidence_diffs else "",
                "candidate_input_tokens": sum(int(row["candidate_input_tokens"]) for row in article_rows),
                "candidate_cached_input_tokens": sum(int(row["candidate_cached_input_tokens"]) for row in article_rows),
                "candidate_output_tokens": sum(int(row["candidate_output_tokens"]) for row in article_rows),
                "candidate_total_tokens": sum(int(row["candidate_total_tokens"]) for row in article_rows),
                "candidate_estimated_cost_usd": round(sum(total_costs), 6) if total_costs else "",
                "disagreements": len(benchmark.disagreement_rows),
            }
        )
    return rows


def write_benchmark_reports(benchmarks: list[ModelBenchmark], out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    summary_rows = summarize_benchmarks(benchmarks)
    article_rows = [row for benchmark in benchmarks for row in benchmark.article_rows]
    disagreement_rows = [row for benchmark in benchmarks for row in benchmark.disagreement_rows]

    _write_csv(out_dir / "benchmark_summary.csv", summary_rows)
    _write_csv(out_dir / "benchmark_articles.csv", article_rows)
    _write_csv(out_dir / "benchmark_disagreements.csv", disagreement_rows)
    _write_xlsx(out_dir / "benchmark.xlsx", summary_rows, article_rows, disagreement_rows)


def safe_model_dir_name(model: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", model).strip("_") or "model"


def _empty_model_summary(model: str) -> dict[str, Any]:
    return {
        "model": model,
        "articles": 0,
        "screening_matches": 0,
        "screening_agreement_rate": "",
        "compared_items": 0,
        "matching_items": 0,
        "answer_agreement_rate": "",
        "mean_confidence_abs_diff": "",
        "candidate_input_tokens": 0,
        "candidate_cached_input_tokens": 0,
        "candidate_output_tokens": 0,
        "candidate_total_tokens": 0,
        "candidate_estimated_cost_usd": "",
        "disagreements": 0,
    }


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    columns = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(
    path: Path,
    summary_rows: list[dict[str, Any]],
    article_rows: list[dict[str, Any]],
    disagreement_rows: list[dict[str, Any]],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to write XLSX benchmark reports.") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in [
        ("Summary", summary_rows),
        ("Articles", article_rows),
        ("Disagreements", disagreement_rows),
    ]:
        sheet = workbook.create_sheet(name)
        columns = list(rows[0].keys()) if rows else []
        if columns:
            sheet.append(columns)
            for row in rows:
                sheet.append([row.get(column, "") for column in columns])
        _style_sheet(sheet, columns, get_column_letter, Font, PatternFill, Alignment)
    workbook.save(path)


def _style_sheet(sheet: Any, columns: list[str], get_column_letter: Any, Font: Any, PatternFill: Any, Alignment: Any) -> None:
    if not columns:
        return
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = _column_width(column)
        for cell in sheet[letter][1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=column.endswith("value") or column == "article_id")
            if column.endswith("rate") or column.endswith("diff"):
                cell.number_format = "0.0000"
            if column.endswith("cost_usd"):
                cell.number_format = "$0.000000"


def _column_width(column: str) -> int:
    if column == "article_id":
        return 38
    if column.endswith("value"):
        return 42
    if column.endswith("tokens"):
        return 18
    if column.endswith("cost_usd"):
        return 20
    return min(max(len(column) + 4, 14), 28)


def _answer_key(answer: FinalAnswer) -> tuple[str, ...]:
    value = answer.final_answer
    if isinstance(value, list):
        return tuple(sorted(str(item).strip().lower() for item in value))
    return (str(value).strip().lower(),)


def _answer_text(answer: FinalAnswer | None) -> str:
    if answer is None:
        return "<missing>"
    value = answer.final_answer
    if isinstance(value, list):
        return "; ".join(str(item) for item in value)
    return str(value)
